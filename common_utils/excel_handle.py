# -*- coding: utf-8 -*-
# @Version: Python 3.9
# @Time    : 2023/1/9 16:42
# @Author  : chenyinhua
# @File    : handle_excel.py
# @Software: PyCharm
# @Desc: 使用openpyxl对excel进行读写操作

import openpyxl
from loguru import logger


class ReadExcel:

    def __init__(self, filename):
        """
        初始化用例文件
        :param filename: 文件绝对路径，如：D:\test\test.xlsx
        """
        self.filename = filename

    def create_excel(self):
        """
        创建excel文件，需要指定excel文件的绝对路径，如D:\test\test.xlsx
        """
        # 创建文件对象
        wb = openpyxl.Workbook()
        # 创建excel文件
        wb.save(self.filename)
        return self.filename

    def read(self, sheet=None):
        """
        读取excel数据并返回
        :param sheet: 表单名称
        :return: 存在传入的表单, 返回表单数据，不存在则返回空
        """
        # 创建一个工作簿工作对象(excel文件已存在的情况)
        workbook = openpyxl.open(self.filename)
        # 跟上面那句一个意思 workbook = openpyxl.load_workbook(self.file)

        # 获取excel当中所有的sheet，返回的是一个列表
        sheets = workbook.sheetnames

        # 如果sheet不为空，则取sheet等于指定sheet
        if sheet:
            sheet = sheet
        # 如果sheet为空，则sheet为第一个表单
        else:
            sheet = sheets[0]

        data = []

        if sheet in sheets:
            sheet = workbook[sheet]
            all_values = list(sheet.values)
            header = all_values[0]
            for i in all_values[1:]:
                data.append(dict(zip(header, i)))
            # 关闭excel
            workbook.close()
            return data
        else:
            # 关闭excel
            workbook.close()
            logger.error(f"表单： 【{sheet}】文件不存在")

    def write(self, row, column, data, sheet_name=None):
        """
        往excel写入数据
        :param sheet_name: 表单名称
        :param row: 要写入的行
        :param column: 要写入的列
        :param data: 要写入的数据
        :return: None
        """
        workbook = openpyxl.open(self.filename)
        # 获取excel当中所有的sheet，返回的是一个列表
        sheets = workbook.sheetnames
        if sheet_name in sheets:
            sheet = workbook[sheet_name]
            logger.info(f"往表单【{sheet_name}】中写入数据")
        else:
            # 如果表单为空，就默认使用第一个表单
            sheet = workbook.active
            logger.info(f"表单【{sheet_name}】不存在，默认往第一个表单中写入数据")

        sheet.cell(row=row, column=column, value=data)
        # 更上面写法效果一样 sheet.cell(row=row, column=column).value = data

        # 保存并关闭文件
        workbook.save(self.filename)
        workbook.close()
